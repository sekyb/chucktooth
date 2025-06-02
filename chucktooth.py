import os
import re
import base64
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.drawing.image import Image as ExcelImage
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
import argparse
import matplotlib.pyplot as plt
import numpy as np
import tempfile
from collections import defaultdict
import logging
from tqdm import tqdm

# --- Logging Setup: File only, no console output ---
LOG_FILE = "Logjam"
logger = logging.getLogger()
logger.setLevel(logging.DEBUG)
fh = logging.FileHandler(LOG_FILE, mode='w')
fh.setLevel(logging.DEBUG)
fh_formatter = logging.Formatter(
    '%(asctime)s [%(levelname)s] %(module)s::%(funcName)s: %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
fh.setFormatter(fh_formatter)
logger.handlers = []
logger.addHandler(fh)

# --- Constants ---
PLACEHOLDER_VALUES = {
    "password", "token", "secret", "string", "null", "none", "default", "example", "changeme"
}
EXT_LANG_MAP = {
    '.py': 'Python', '.js': 'JavaScript', '.html': 'HTML', '.css': 'CSS',
    '.json': 'JSON', '.yml': 'YAML', '.yaml': 'YAML', '.ini': 'INI', '.cfg': 'CFG',
    '.c': 'C', '.cpp': 'C++', '.cc': 'C++', '.h': 'C/C++ Header', '.hpp': 'C++ Header',
    '.java': 'Java', '.rb': 'Ruby', '.go': 'Go', '.php': 'PHP', '.sh': 'Shell',
    '.bat': 'Batch', '.pl': 'Perl', '.swift': 'Swift', '.kt': 'Kotlin', '.ts': 'TypeScript'
}

PATTERNS = {
    "IP Address (IPv4)": r"\b(?:(?:25[0-5]|2[0-4]\d|1\d{2}|[1-9]?\d)\.){3}(?:25[0-5]|2[0-4]\d|1\d{2}|[1-9]?\d)\b",
    "IP Address (IPv6)": r"\b(?:[A-Fa-f0-9]{1,4}:){2,7}[A-Fa-f0-9]{1,4}\b|\b::(?:[A-Fa-f0-9]{1,4}:){0,5}[A-Fa-f0-9]{1,4}\b",
    "Email": r"\b[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,7}\b",
    "Password/Token/Secret": r"""(?ix)
        \b(password|secret|pass(word)?|passwd|api_?key|token|access_token|client_secret|client_id)\b
        \s* (?:=|:)\s* 
        ['"]
        (?P<secret>[A-Za-z0-9!@#$%^&*()_+={}\[\]|;:<>,.?/~`-]{8,})
        ['"]
    """,
    "URL": r"\b(?:https?|ftp|file)://[^\s\"'><)]+",
    "Base64 Encoded": r"\b(?:[A-Za-z0-9+/]{40,}={0,2})\b",
    "Credit Card": r"\b(?:4[0-9]{12}(?:[0-9]{3})?|5[1-5][0-9]{14}|3[47][0-9]{13}|6(?:011|5[0-9]{2})[0-9]{12})\b",
    "AWS Access Key": r"\bAKIA[0-9A-Z]{16}\b",
    "AWS Secret Key": r"\b(?:ASIA|AKIA)[A-Za-z0-9/+=]{16,40}\b",
    "Private Key": r"-----BEGIN (RSA |EC |DSA )?PRIVATE KEY-----[\s\S]+?-----END (RSA |EC |DSA )?PRIVATE KEY-----",
    "MongoDB URI": r"\bmongodb(?:\+srv)?:\/\/[^\s\"'<>]+",
    "Docker Credentials": r"//([A-Za-z0-9._-]+:[A-Za-z0-9!@#$%^&*()_+={}|;:<>,.?/~`-]{8,})@",
    "Generic Token": r"\b(?=[a-zA-Z0-9-_]{32,64}\b)(?=.*[a-zA-Z])(?=.*\d)[a-zA-Z0-9-_]{32,64}\b",
    "Ethereum Address": r"\b0x[a-fA-F0-9]{40}\b",
    "Bitcoin Address": r"\b([13][a-km-zA-HJ-NP-Z1-9]{25,34}|bc1[a-zA-HJ-NP-Z0-9]{11,71})\b",
}
PATTERN_FLAGS = {
    "Password/Token/Secret": re.IGNORECASE | re.VERBOSE,
    "Private Key": re.DOTALL
}

def is_potential_secret(s):
    if not isinstance(s, str): return False
    s_lower = s.lower()
    if s_lower in PLACEHOLDER_VALUES: return False
    if len(set(s)) <= 3: return False
    if s.isdigit() or s.islower() or s.isupper(): return False
    if len(s) < 8: return False
    return True

def remove_illegal_excel_chars(val):
    if isinstance(val, str):
        return re.sub(r"[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F]", "", val)
    return val

def safe_base64_decode(data):
    try:
        clean_data = ''.join(data.split())
        missing_padding = len(clean_data) % 4
        if missing_padding:
            clean_data += '=' * (4 - missing_padding)
        decoded = base64.b64decode(clean_data, validate=False)
        try:
            return decoded.decode("utf-8")
        except Exception:
            return repr(decoded)
    except Exception:
        return ""

def scan_file(file_path):
    logger.debug(f"Scanning file: {file_path}")
    results = []
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            for i, line in enumerate(f, start=1):
                for key, pattern in PATTERNS.items():
                    flags = PATTERN_FLAGS.get(key, 0)
                    for match in re.finditer(pattern, line, flags):
                        decoded_val = ""
                        match_text = match.group(0)
                        if key == "Password/Token/Secret":
                            secret = match.group("secret")
                            if not is_potential_secret(secret): continue
                            match_text = secret
                        elif key == "Docker Credentials":
                            match_text = match.group(1)
                            if not is_potential_secret(match_text.split(':', 1)[-1]): continue
                        elif key == "Private Key":
                            continue
                        elif key == "Base64 Encoded":
                            decoded_val = safe_base64_decode(match_text)
                        results.append({
                            'File': file_path,
                            'Line': i,
                            'Pattern': key,
                            'Match': match_text,
                            'Decoded': decoded_val
                        })
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
            priv_key_pattern = PATTERNS["Private Key"]
            for match in re.finditer(priv_key_pattern, content, re.DOTALL):
                start_line = content[:match.start()].count('\n') + 1
                match_text = match.group(0)
                logger.warning(f"Private Key found in {file_path} (line {start_line})")
                results.append({
                    'File': file_path,
                    'Line': start_line,
                    'Pattern': "Private Key",
                    'Match': match_text[:60] + ('...' if len(match_text) > 60 else ''),
                    'Decoded': ""
                })
    except Exception as e:
        logger.error(f"Error scanning file {file_path}: {e}")
    return results

def scan_directory(directory):
    logger.info(f"Scanning directory: {directory}")
    all_files = []
    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith(('.txt', '.py', '.js', '.html', '.json', '.env', '.yml', '.yaml', '.ini', '.cfg')) or '.' in file:
                all_files.append(os.path.join(root, file))
    logger.info(f"Total files to scan: {len(all_files)}")
    results = []
    with ThreadPoolExecutor() as executor:
        futures = [executor.submit(scan_file, file_path) for file_path in all_files]
        for f in tqdm(as_completed(futures), total=len(futures), desc="Scanning files"):
            results.extend(f.result())
    logger.info(f"Total findings: {len(results)}")
    return results

def count_lines_per_language(directory):
    logger.info(f"Counting lines of code in directory: {directory}")
    stats = defaultdict(int)
    for root, dirs, files in os.walk(directory):
        for file in files:
            ext = os.path.splitext(file)[1].lower()
            lang = EXT_LANG_MAP.get(ext)
            if lang:
                try:
                    with open(os.path.join(root, file), 'r', encoding='utf-8', errors='ignore') as f:
                        line_count = sum(1 for _ in f)
                        stats[lang] += line_count
                except Exception as e:
                    logger.error(f"Error reading file {file}: {e}")
    if stats:
        logger.info(f"Language stats: {dict(stats)}")
        df = pd.DataFrame(list(stats.items()), columns=["Language", "Lines of Code"])
        return df
    else:
        logger.warning("No lines of code counted.")
        return pd.DataFrame(columns=["Language", "Lines of Code"])

def plot_loc_chart(df, img_path):
    if df is not None and not df.empty:
        logger.info(f"Generating LOC bar chart: {img_path}")
        plt.figure(figsize=(8, 5))
        plt.bar(df['Language'], df['Lines of Code'])
        plt.xticks(rotation=45, ha="right")
        plt.ylabel("Lines of Code")
        plt.title("Lines of Code by Language")
        plt.tight_layout()
        plt.savefig(img_path)
        plt.close()

def plot_findings_chart(results, img_path):
    df = pd.DataFrame(results)
    if not df.empty and 'Pattern' in df:
        logger.info(f"Generating findings bar chart: {img_path}")
        count = df['Pattern'].value_counts()
        plt.figure(figsize=(8, 5))
        count.plot(kind="bar")
        plt.ylabel("Occurrences")
        plt.title("Sensitive Findings by Category")
        plt.tight_layout()
        plt.savefig(img_path)
        plt.close()

def plot_pattern_summary_pie_chart(results, img_path):
    df = pd.DataFrame(results)
    if not df.empty and 'Pattern' in df:
        logger.info(f"Generating pattern summary pie chart: {img_path}")
        count = df['Pattern'].value_counts()
        labels = count.index.tolist()
        sizes = count.values
        total = np.sum(sizes)
        legend_labels = [f"{label} ({size} / {size/total:.1%})" for label, size in zip(labels, sizes)]
        plt.figure(figsize=(8, 5))
        wedges, texts = plt.pie(
            sizes,
            labels=None,
            startangle=140
        )
        plt.title("Findings by Pattern Type")
        plt.legend(wedges, legend_labels, title="Pattern", loc="center left", bbox_to_anchor=(1, 0.5))
        plt.tight_layout(rect=[0, 0, 0.8, 1])  # leave space for legend
        plt.savefig(img_path, bbox_inches='tight')
        plt.close()

def save_to_excel(results, output_file, loc_df=None, loc_img=None, findings_img=None, pattern_summary_img=None):
    logger.info(f"Writing results to Excel file: {output_file}")
    df = pd.DataFrame(results)
    if "Decoded" not in df.columns:
        df["Decoded"] = ""
    for col in df.columns:
        df[col] = df[col].map(remove_illegal_excel_chars)
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Tab 1: Exec Summary
        if loc_df is not None and not loc_df.empty:
            loc_df.to_excel(writer, index=False, sheet_name='Exec Summary', startrow=1)
            workbook = writer.book
            worksheet = writer.sheets['Exec Summary']
            banner_text = f"CONFIDENTIAL // sekyb // {datetime.now().year}"
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(loc_df.columns))
            cell = worksheet.cell(row=1, column=1)
            cell.value = banner_text
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True, color="FFFFFF")
            worksheet.column_dimensions['A'].width = 22
            worksheet.column_dimensions['B'].width = 18
            img_offset_row = len(loc_df)+4
            if loc_img and os.path.exists(loc_img):
                try:
                    img = ExcelImage(loc_img)
                    img.anchor = f"A{img_offset_row}"
                    worksheet.add_image(img)
                    img_offset_row += 25  # More space between images
                except Exception as e:
                    logger.error(f"Error inserting LOC chart in Excel: {e}")
            if pattern_summary_img and os.path.exists(pattern_summary_img):
                try:
                    img2 = ExcelImage(pattern_summary_img)
                    img2.anchor = f"A{img_offset_row}"
                    worksheet.add_image(img2)
                except Exception as e:
                    logger.error(f"Error inserting pattern summary chart in Excel: {e}")
        # Tab 2: Results (Sensitive Findings)
        df.to_excel(writer, index=False, sheet_name='Results', startrow=2)
        worksheet2 = writer.sheets['Results']
        banner_text = f"CONFIDENTIAL // sekyb // {datetime.now().year}"
        worksheet2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        cell = worksheet2.cell(row=1, column=1)
        cell.value = banner_text
        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True, color="FFFFFF")
        header_row = 3
        for idx, col in enumerate(df.columns, 1):
            cell = worksheet2.cell(row=header_row, column=idx)
            cell.font = Font(bold=True)
            max_length = max(df[col].astype(str).map(len).max(), len(str(col))) + 2
            if col == "File":
                max_length = min(max_length, 50)
            elif col in ("Match", "Decoded"):
                max_length = min(max_length, 100)  # Allow wider columns for Match and Decoded
            worksheet2.column_dimensions[get_column_letter(idx)].width = max_length
            for row in range(header_row + 1, header_row + 1 + len(df)):
                alignment = Alignment(horizontal="left") if col == "Line" else Alignment(horizontal="general")
                worksheet2.cell(row=row, column=idx).alignment = alignment
        last_col_letter = get_column_letter(len(df.columns))
        worksheet2.auto_filter.ref = f"A{header_row}:{last_col_letter}{header_row}"
        worksheet2.freeze_panes = worksheet2[f"A{header_row + 1}"]
        if findings_img and os.path.exists(findings_img):
            try:
                img2 = ExcelImage(findings_img)
                img2.anchor = f"A{header_row + len(df) + 3}"
                worksheet2.add_image(img2)
            except Exception as e:
                logger.error(f"Error inserting findings chart in Excel: {e}")
    logger.info("Excel report generation complete.")

def main():
    parser = argparse.ArgumentParser(description="Scan a directory for sensitive data and code statistics.")
    parser.add_argument('-d', '--directory', required=True, help='Directory to scan')
    parser.add_argument('-o', '--output', default='results.xlsx', help='Output Excel file name (default: results.xlsx)')
    args = parser.parse_args()
    output_file = args.output
    if not output_file.endswith('.xlsx'):
        output_file += '.xlsx'
    logger.info("STARTING SCAN")
    logger.info(f"Parameters: directory={args.directory}, output={output_file}")
    print(f"Scanning directory {args.directory}...")
    results = scan_directory(args.directory)
    print("Counting lines of code by language...")
    loc_df = count_lines_per_language(args.directory)
    with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as loc_img:
        loc_img_path = loc_img.name
    if loc_df is not None and not loc_df.empty:
        plot_loc_chart(loc_df, loc_img_path)
    else:
        loc_img_path = None
    with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as findings_img:
        findings_img_path = findings_img.name
    if results:
        plot_findings_chart(results, findings_img_path)
    else:
        findings_img_path = None
    with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as pattern_summary_img:
        pattern_summary_img_path = pattern_summary_img.name
    if results:
        plot_pattern_summary_pie_chart(results, pattern_summary_img_path)
    else:
        pattern_summary_img_path = None
    print(f"Saving results to {output_file}...")
    save_to_excel(
        results, output_file,
        loc_df=loc_df,
        loc_img=loc_img_path,
        findings_img=findings_img_path,
        pattern_summary_img=pattern_summary_img_path
    )
    logger.info("PROCESS COMPLETE")
    print("Scan complete and results saved!")

if __name__ == "__main__":
    main()